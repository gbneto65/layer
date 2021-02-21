# Layers performance project - branch V2
# module: create a CSV database from Excel file (original data)
# 
# to be prepare the databank for the layers performance system
# 
# create a new CSV file everytime the database from de layers was updated or new layer genetics was inputed.
# as the CSV file is the file that will be load into the system.
# do not create any new sheet without the same strucutre of the others.
# the sheet_name should be the name of the genetic line of the layers
#
#Feb. 20th 2021
#
#setup parameters are defined in other file: setup-parameter.py
#excel_file = 'layers_dta.xlsx' # name of excel file that contains the layers genetics performance data
#excel_col_import = 'A:N' # define the coluns that will be imported from the databank

import pandas as pd
import os
from operator import itemgetter
from openpyxl import load_workbook
import winsound
from setup_parameters_v2 import excel_file, excel_col_import, csv_file_name, working_directory, sound_error


# get the spreadsheets names and return as a list
def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only = True, keep_links = False)
    return wb.sheetnames

print(f'working directory {working_directory}')
print(f'\nExcel file Name : {excel_file}')
print(f'Excel coluns to import : {excel_col_import}')
print(f'CSV file to be created : {csv_file_name}\n')

# read excel file to get the spreadsheets names (names of the genetics)
try:
	genetics_list = get_sheetnames_xlsx(excel_file)
	print('XLSX file found\n')
	print(f'Found data from : {genetics_list}\n')
except :
	print ('could not find XLSX file')
	winsound.Beep(sd1[0],sd1[1])
	winsound.Beep(sd2[0],sd2[1])

# generate a list from the genetics
dta = []
for i in range (len(genetics_list)):
	dta.append(i)


for i in range (len(genetics_list)):
	dta[i] = pd.read_excel(excel_file,
		sheet_name=genetics_list[i],
		usecols=excel_col_import,
		)

# concatenate the dataframes
layers_dta = pd.concat(dta, axis=0)

# create CSV file
if csv_file_name == '':
	print('Error - verify the CSV name of the file at - SETUP_PARAMETERS_v2.PY')
	sound_error()
else:
	layers_dta.to_csv(csv_file_name)
	print('CSV file successfully created')
	
#print(layers_dta)
print('\nDatabase created with sucess! \nDone!!!')


